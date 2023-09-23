import googlemaps
import folium
from openpyxl import load_workbook


def transform_coordinates(coord_str):
    print(f"Transforming: {coord_str}")
    hemisphere = coord_str[-1]  # 'N' or 'S' for latitude, 'E' or 'W' for longitude

    # Splitting the string to separate degrees, minutes, and seconds
    coord_parts = coord_str[:-1].split('N') if 'N' in coord_str else coord_str[:-1].split('E')
    degrees = int(coord_parts[0])
    minutes_seconds = coord_parts[1]
    minutes = int(minutes_seconds[:2])
    seconds = float(minutes_seconds[2:])

    # Normalize minutes and seconds
    while seconds >= 60.0:
        seconds -= 60.0
        minutes += 1
    while minutes >= 60:
        minutes -= 60
        degrees += 1

    formatted_coord = f"{degrees}º {minutes:02d}' {seconds:05.2f}''{hemisphere}"
    print(f"Transformed to: {formatted_coord}")
    return formatted_coord, degrees, minutes, seconds, hemisphere


def convert_to_decimal(degrees, minutes, seconds, hemisphere):
    decimal = degrees + minutes / 60.0 + seconds / 3600.0
    if hemisphere in ['S', 'W']:
        decimal *= -1
    return decimal


# Configure Google Maps API
gmaps = googlemaps.Client(key='AIzaSyCuqRvgqRQe0zUwtsZ_sq8-oIM3GP4XCkU')

# Initialize a folium map centered around a starting coordinate
m = folium.Map(location=[41.7266, 1.8261], zoom_start=10)

# Load Excel workbook
wb = load_workbook('coords.xlsx')
ws = wb.active

# Loop through rows to update them
for row in ws.iter_rows(min_row=2, min_col=1, max_col=2):
    lat_cell, lon_cell = row
    print(f"Processing row {lat_cell.row}")
    print(f"Latitude cell value: {lat_cell.value}")
    print(f"Longitude cell value: {lon_cell.value}")

    formatted_lat, lat_deg, lat_min, lat_sec, lat_hem = transform_coordinates(lat_cell.value)
    formatted_lon, lon_deg, lon_min, lon_sec, lon_hem = transform_coordinates(lon_cell.value)

    lat_decimal = convert_to_decimal(lat_deg, lat_min, lat_sec, lat_hem)
    lon_decimal = convert_to_decimal(lon_deg, lon_min, lon_sec, lon_hem)

    # Use reverse geocoding to get the municipality
    reverse_geocode_result = gmaps.reverse_geocode((lat_decimal, lon_decimal))
    municipality = reverse_geocode_result[0]['address_components'][2]['long_name']

    ws.cell(row=lat_cell.row, column=3).value = municipality

    # Add marker to folium map
    folium.Marker([lat_decimal, lon_decimal]).add_to(m)

# Save changes to Excel
wb.save('updated_coords.xlsx')

# Save the folium map to HTML
m.save('map.html')
